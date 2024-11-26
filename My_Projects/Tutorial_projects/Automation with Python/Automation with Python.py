import openpyxl as xl

from openpyxl.chart import BarChart,Reference # We imported two classes BarChart and Reference
# adds an alias for openpyxl

def process_workbook(filename):
    workbook = xl.load_workbook(filename)
    sheet = workbook['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row=row, column=4)
        corrected_price_cell.value = corrected_price

    Values = Reference(sheet, 
            min_row=2 , 
            max_row = sheet.max_row,
            min_col=4,
            max_col=4)

    Chart = BarChart()
    Chart.add_data(Values)
    sheet.add_chart(Chart, 'A7')

    workbook.save(filename)

process_workbook('lol.xlsx')